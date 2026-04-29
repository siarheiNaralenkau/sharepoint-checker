from __future__ import annotations

import logging
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from ..models.result_models import RunSummary, CheckStatus

logger = logging.getLogger(__name__)

_HEADERS = [
    "Site Name",
    "Leadership Folder",
    "Status",
    "Failure Reason",
    "Site URL",
    "Roaster Folder",
    "Roaster has files",
    "File modification date",
    "Generated Time",
]

_FILL_PASS = PatternFill(fill_type="solid", fgColor="C6EFCE")   # Excel green
_FILL_FAIL = PatternFill(fill_type="solid", fgColor="FFC7CE")   # Excel red
_FONT_HEADER = Font(bold=True, color="FFFFFF")
_FILL_HEADER = PatternFill(fill_type="solid", fgColor="0078D4")  # Microsoft blue
_ALIGN_WRAP = Alignment(wrap_text=True, vertical="top")


def write_xlsx_report(summary: RunSummary, output_dir: str | Path) -> Path:
    out = Path(output_dir)
    out.mkdir(parents=True, exist_ok=True)
    local_dt = summary.started_at.astimezone()
    date_str = local_dt.strftime("%d-%m-%Y_%H-%M-%S")
    generated_time = local_dt.strftime("%d/%m/%Y %H:%M:%S")
    path = out / f"SAP SE Account - Roaster Review - {date_str}.xlsx"

    sites = [s for s in summary.site_results if s.leadership_folder is not None]

    wb = Workbook()
    ws = wb.active
    ws.title = "Site Results"

    # Header row
    ws.append(_HEADERS)
    for col_idx in range(1, len(_HEADERS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = _FONT_HEADER
        cell.fill = _FILL_HEADER
        cell.alignment = _ALIGN_WRAP

    # Data rows
    _SITE_URL_COL = _HEADERS.index("Site URL") + 1

    for site in sites:
        row = [
            site.report_display_name,
            site.leadership_folder or "",
            site.overall_status.value,
            site.failure_reason or "",
            site.site_url or "",
            site.roaster_found,
            site.roaster_has_files,
            site.roaster_last_modified or "",
            generated_time,
        ]
        ws.append(row)

        fill = _FILL_PASS if site.overall_status == CheckStatus.PASS else _FILL_FAIL
        row_idx = ws.max_row
        for col_idx in range(1, len(_HEADERS) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = fill
            cell.alignment = _ALIGN_WRAP

        # Make site_url cell a clickable hyperlink
        if site.site_url:
            url_cell = ws.cell(row=row_idx, column=_SITE_URL_COL)
            url_cell.hyperlink = site.site_url
            url_cell.font = Font(color="0078D4", underline="single")

    # Column widths
    col_widths = [50, 35, 10, 55, 60, 16, 18, 22, 22]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"

    wb.save(path)
    logger.info("XLSX report written to %s (%d site(s) after filtering)", path, len(sites))
    return path
