import json
import tempfile
from datetime import datetime, timezone

import pytest
from openpyxl import load_workbook

from sharepoint_checker.models.result_models import (
    CheckStatus,
    RunSummary,
    SiteCheckResult,
)
from sharepoint_checker.reporting import write_json_report, write_xlsx_report, write_html_report


def _make_summary(fail: bool = False) -> RunSummary:
    # Fail scenario: leadership folder exists but Roster is missing (stays in report after filtering)
    site = SiteCheckResult(
        site_name="EPAMSAPSEProjectsCSDArea",
        site_url="https://epam.sharepoint.com/sites/EPAMSAPSEProjectsCSDArea",
        site_id="epam.sharepoint.com,abc,def",
        display_name="EPAM SAP SE Projects, CSD Area-Project SAP-MxG leadership",
        drive_id="drive1",
        leadership_folder="Project SAP-MxG leadership",
        roster_found=not fail,
        roster_has_files=not fail,
        failure_reason=None if not fail else "'Roster' folder not found inside 'Project SAP-MxG leadership'",
        overall_status=CheckStatus.FAIL if fail else CheckStatus.PASS,
    )
    return RunSummary(
        run_id="2026-04-20T12:00:00Z",
        started_at=datetime(2026, 4, 20, 12, 0, 0, tzinfo=timezone.utc),
        completed_at=datetime(2026, 4, 20, 12, 1, 0, tzinfo=timezone.utc),
        site_results=[site],
        total_sites=1,
        pass_count=0 if fail else 1,
        fail_count=1 if fail else 0,
        overall_status=CheckStatus.FAIL if fail else CheckStatus.PASS,
    )


def test_json_report_structure():
    summary = _make_summary()
    with tempfile.TemporaryDirectory() as d:
        path = write_json_report(summary, d)
        data = json.loads(path.read_text())

    # run_id renamed to reporting_datetime and moved to last key
    assert "run_id" not in data
    assert data["reporting_datetime"] == "2026-04-20T12:00:00Z"
    assert list(data.keys())[-1] == "reporting_datetime"

    assert len(data["site_results"]) == 1
    site = data["site_results"][0]
    keys = list(site.keys())
    assert keys[0] == "display_name"
    assert keys[1] == "status"
    assert site["status"] == "PASS"
    assert site["site_url"] == "https://epam.sharepoint.com/sites/EPAMSAPSEProjectsCSDArea"
    assert site["roaster_folder"] is True
    assert site["roaster_has_files"] is True
    assert site["failure_reason"] == ""
    assert site["reporting_datetime"] == "2026-04-20T12:00:00Z"
    # internal fields must not leak
    assert "site_id" not in site
    assert "site_name" not in site
    assert "overall_status" not in site
    assert "roster_found" not in site


def test_json_report_filters_null_leadership():
    """Sites without a leadership_folder must be excluded from the report."""
    site_no_folder = SiteCheckResult(
        site_name="NoFolder",
        site_url="https://epam.sharepoint.com/sites/NoFolder",
        site_id="site-no-folder",
        display_name="No Leadership Site",
        leadership_folder=None,
        overall_status=CheckStatus.FAIL,
        failure_reason="No folder matching regex found at root",
    )
    summary = _make_summary()
    summary.site_results.append(site_no_folder)
    with tempfile.TemporaryDirectory() as d:
        path = write_json_report(summary, d)
        data = json.loads(path.read_text())
    assert len(data["site_results"]) == 1  # filtered site excluded
    assert all(s["leadership_folder"] is not None for s in data["site_results"])


def test_json_report_fail_site():
    summary = _make_summary(fail=True)
    with tempfile.TemporaryDirectory() as d:
        path = write_json_report(summary, d)
        data = json.loads(path.read_text())
    site = data["site_results"][0]
    assert site["status"] == "FAIL"
    assert site["roaster_folder"] is False
    assert "Roster" in site["failure_reason"]


def test_xlsx_report_rows_and_colors():
    summary = _make_summary(fail=True)
    with tempfile.TemporaryDirectory() as d:
        path = write_xlsx_report(summary, d)
        wb = load_workbook(path)
        ws = wb.active

    headers = [ws.cell(row=1, column=i).value for i in range(1, 9)]
    # display_name first, status second, site_url third, reporting_datetime last
    assert headers[0] == "display_name"
    assert headers[1] == "status"
    assert headers[2] == "site_url"
    assert headers[-1] == "reporting_datetime"
    assert "roaster_folder" in headers
    assert "site_id" not in headers

    data_row = [ws.cell(row=2, column=i).value for i in range(1, 9)]
    assert "FAIL" in data_row

    # FAIL row must have red fill (openpyxl reads 6-char hex back with 00 alpha prefix)
    fill_color = ws.cell(row=2, column=1).fill.fgColor.rgb
    assert fill_color == "00FFC7CE"


def test_xlsx_report_pass_row_green():
    summary = _make_summary(fail=False)
    with tempfile.TemporaryDirectory() as d:
        path = write_xlsx_report(summary, d)
        wb = load_workbook(path)
        ws = wb.active

    fill_color = ws.cell(row=2, column=1).fill.fgColor.rgb
    assert fill_color == "00C6EFCE"


def test_xlsx_report_filters_null_leadership():
    site_no_folder = SiteCheckResult(
        site_name="NoFolder",
        site_url="https://epam.sharepoint.com/sites/NoFolder",
        site_id="site-no-folder",
        display_name="No Leadership Site",
        leadership_folder=None,
        overall_status=CheckStatus.FAIL,
        failure_reason="No folder matching regex found at root",
    )
    summary = _make_summary()
    summary.site_results.append(site_no_folder)
    with tempfile.TemporaryDirectory() as d:
        path = write_xlsx_report(summary, d)
        wb = load_workbook(path)
        ws = wb.active
    assert ws.max_row == 2  # header + 1 data row (filtered site excluded)


def test_html_report_contains_key_data():
    summary = _make_summary(fail=True)
    with tempfile.TemporaryDirectory() as d:
        path = write_html_report(summary, d)
        html = path.read_text(encoding="utf-8")
    assert "2026-04-20T12:00:00Z" in html
    assert "EPAM SAP SE Projects, CSD Area-Project SAP-MxG leadership" in html
    assert "FAIL" in html
    assert "fail-row" in html
    assert "Reporting DateTime" in html
    assert "run_id" not in html


def test_html_report_filters_null_leadership():
    site_no_folder = SiteCheckResult(
        site_name="NoFolder",
        site_url="https://epam.sharepoint.com/sites/NoFolder",
        site_id="site-no-folder",
        display_name="No Leadership Site",
        leadership_folder=None,
        overall_status=CheckStatus.FAIL,
        failure_reason="No folder matching regex found at root",
    )
    summary = _make_summary()
    summary.site_results.append(site_no_folder)
    with tempfile.TemporaryDirectory() as d:
        path = write_html_report(summary, d)
        html = path.read_text(encoding="utf-8")
    assert "No Leadership Site" not in html
